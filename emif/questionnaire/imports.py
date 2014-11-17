# Copyright (C) 2014 Ricardo Ribeiro and Universidade de Aveiro
#
# Authors: Ricardo Ribeiro <ribeiro.r@ua.pt>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#

from questionnaire.models import Questionnaire, Choice, Question, QuestionSet

import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Style, PatternFill, Alignment, Font, Border, Side
from openpyxl.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation, ValidationType

from searchengine.search_indexes import convert_text_to_slug
from searchengine.models import Slugs

from questionnaire.utils import *

import datetime

from django.db import transaction

"""This class is used to import the fingerprint template
"""
class ImportQuestionnaire(object):


    def __init__(self, file_path):
        self.file_path = file_path

    def import_questionnaire(self):
        raise NotImplementedError("Please Implement this method")

    def writeLog(self, log):
        pass
        #with open("log_%s.txt" % datetime.datetime.now().strftime("%Y%m%d-%H%M%S"), "w") as f:
        #    f.write(log)
        #    f.close()

    def get_slug(self, slug, questionnaire):
        return next_free_slug(slug, create=False, scope=questionnaire)
        # slugs_objs = Slugs.objects.filter(slug1=slug)
        # slug_aux = None
        # if (len(slugs_objs)>0):
        #     slug_aux=slugs_objs[0]
        # else:
        #     return slug
        # slug_name_final = ""
        # slug_arr = slug_aux.slug1.split("_")
        # if len(slug_arr)>0:
        #     if (slug_arr[len(slug_arr)-1].isdigit()):
        #         slug_number = int(slug_arr[len(slug_arr)-1]) + 1
        #         slug_arr[len(slug_arr)-1] = str(slug_number)
        #         slug_name_final = "_".join(slug_arr)

        #     else:
        #         slug_name_final = slug_aux.slug1 + "_0"
        # else:
        #     slug_name_final = slug_aux.slug1 + "_0"

        # return slug_name_final

    def format_number(self, number):
        # print number
        number_arr = number.split(".")

        result = number_arr[0] + "."
        for i in range(1,len(number_arr)):

            val = int(number_arr[i])
            if val<10:
                val = "0" + str(val)
            number_arr[i] = str(val)
            if (i!=len(number_arr)-1):
                result += str(val) + "."
            else:
                result += str(val)
        # print "result " + result
        return result

    """This method will build the object according with the type
    of the object to import.
    """
    @staticmethod
    def factory(t_type, file_path):
        if t_type == "excel":
            return ImportQuestionnaireExcel(file_path)

        else:
            raise Exception("The supplied format is not supported")

class CommentPlaceholder:
    value='comment'

class ImportQuestionnaireExcel(ImportQuestionnaire):

    QUESTION=0
    CATEGORY=1

    def __init__(self, file_path):
        ImportQuestionnaire.__init__(self, file_path)

    def __handleQuestion(self, type, row,type_Column, level_number_column, text_question_Column, _questions_rows,
        _choices_array, qNumber, questionset, log, _checks, _debug, questionnaire):
        try:
            text_en = str(level_number_column.value) + '. ' + str(text_question_Column.value)

            dataType_column = None
            if type == self.CATEGORY:
                dataType_column = CommentPlaceholder()
            else:
                dataType_column = row[3]

            if row[7].value:
                slug = row[7].value
            else:
                slug = convert_text_to_slug(str(row[1].value)[:50])

                slug = self.get_slug(slug, questionnaire)

            if row[5].value:
                helpText = row[5].value
            else:
                helpText = ''

            _tooltip = False

            if row[6].value:
                if str(row[6].value).lower() == 'yes':
                    _tooltip = True

            #If has dependencies
            if row[8].value:
                try:
                    dependencies_list = row[8]
                    list_dep_aux = dependencies_list.value.split('|')
                    question_num_parent = (_questions_rows.get(int(list_dep_aux[0])))


                    index_aux = int(str(list_dep_aux[1]))-1
                    choice_parent_list = _choices_array.get(int(list_dep_aux[0]))
                    choice_parent = choice_parent_list[index_aux]
                    _checks = 'dependent=\"' + str(question_num_parent) + ',' + str(choice_parent) + '\"'
                except:
                    raise

            try:
                questionNumber = qNumber.getNumber(level_number_column.value)
                questionNumber = self.format_number(str(questionNumber))
            except:
                if type==self.QUESTION:
                    log += "\n%s - Error to create question number %s" % (type_Column.row, text_en)
                elif type==self.CATEGORY:
                    log += "\n%s - Error to create Category number %s" % (type_Column.row, text_en)

                self.writeLog(log)
                raise

            #print slug
            #Create or load slug
            slugs = Slugs.objects.filter(slug1=slug, description=text_en)
            if len(slugs) <= 0:
                slug_db = Slugs(slug1=slug, description=text_en)
                slug_db.save()
            else:
                slug_db = slugs[0]

            visible_default = False
            if row[10].value:
                if str(row[10].value).lower() == 'visible':
                    visible_default = True

            is_category=None
            is_stats=None

            if type==self.QUESTION:
                is_stats=True
                is_category=False
            elif type==self.CATEGORY:
                is_stats=False
                is_category=True

            question = Question(questionset=questionset, text_en=text_en, number=str(questionNumber),
                                type=dataType_column.value, help_text=helpText, slug=slug, slug_fk=slug_db, stats=True,
                                category=False, tooltip=_tooltip, checks=_checks, visible_default=visible_default)


            if not _debug:
                question.save()

                if type==self.QUESTION:
                    log += '\n%s - Question created %s ' % (type_Column.row, question)
                elif type==self.CATEGORY:
                    log += '\n%s - Category created %s ' % (type_Column.row, question)



            _questions_rows[type_Column.row] = str(questionNumber)

            if type == self.QUESTION:
                if dataType_column.value in ['choice', 'choice-freeform', 'choice-multiple', 'choice-multiple-freeform']:
                    _choices_array_aux = []
                    # Parse of values list
                    values_list = row[4]
                    if (values_list!=None and values_list.value!=None):
                        list_aux = values_list.value.split('|')
                        i = 1
                        for ch in list_aux:
                            try:
                                choice = Choice(question=question, sortid=i, text_en=ch, value=ch)
                                log += '\n%s - Choice created %s ' % (type_Column.row, choice)
                                if not _debug:
                                    choice.save()
                                _choices_array_aux.append(ch)

                                log += '\n%s - Choice saved %s ' % (type_Column.row, choice)
                                i += 1
                            except:
                                log += "\n%s - Error to save Choice %s" % (type_Column.row, choice)
                                self.writeLog(log)
                                raise
                        _choices_array[type_Column.row] = _choices_array_aux

                if dataType_column.value in ['choice-yesno', 'choice-yesnocomment',
                                                     'choice-yesnodontknow']:
                    _choices_array[type_Column.row] = ['yes', 'no', 'dontknow']

        except:
            log += "\n%s - Error to save question %s" % (type_Column.row, text_en)

            self.writeLog(log)
            raise
    @transaction.commit_on_success
    def import_questionnaire(self):
        _debug = False

        qNumber = QuestionNumber()
        slugs = []

        wb = load_workbook(filename = self.file_path, data_only=True)
        ws = wb.get_active_sheet()
        log = ''

        # Cell B1: Name of questionnaire
        name = ws.cell('B1').value
        slugQ = convert_text_to_slug(ws.cell('B1').value)
        disable = False

        questionnaire = Questionnaire(name=name, disable=disable, slug=slugQ, redirect_url='/')
        log += '\nQuestionnaire created %s ' % questionnaire

        try:
            if not _debug:
                questionnaire.save()
                log += '\nQuestionnaire saved %s ' % questionnaire
            _choices_array = {}
            _questions_rows = {}

            #############################
            # TIPS:
            # Type of Row: QuestionSet, Category, Question
            # Columns: Type, Text/Question, Level/Number, Data type, Value list, Help text/Description, Tooltip, Slug, Stats
            #############################
            for row in ws.rows[2:]:
                if len(row) > 0 and row[0].value != None:
                    type_Column = row[0]

                    text_question_Column = row[1]

                    if (text_question_Column.value!=None):
                        text_question_Column.value = text_question_Column.value.encode('ascii', 'ignore')
                    level_number_column = row[2]
                    _checks = ''

                    # Type = QUESTIONSET
                    # Columns required:  Type, Text/Question
                    # Columns optional:  Help text/Description, Tooltip
                    if str(type_Column.value) == "QuestionSet":

                        sortid = str(level_number_column.value)
                        try:
                            qNumber.getNumber('h0', sortid)
                        except:
                            self.writeLog(log)
                            raise
                        text_en = 'h1. %s' % text_question_Column.value
                        slug_qs = str(slugQ) + "_" + convert_text_to_slug(str(text_question_Column.value))
                        if row[5].value:
                            helpText = row[5].value
                        else:
                            helpText = ""
                        tooltip = False
                        if row[6].value:
                            if str(row[6].value).lower() == 'yes':
                                tooltip = True

                        questionset = QuestionSet(questionnaire=questionnaire, checks='required', sortid=sortid, text_en=text_en, heading=slug_qs, help_text=helpText, tooltip=tooltip)
                        log += '\n%s - QuestionSet created %s - %s ' % (type_Column.row, sortid, text_en)
                        try:
                            if not _debug:
                                questionset.save()
                                log += '\n%s - QuestionSet saved %s - %s ' % (type_Column.row, sortid, text_en)
                        except:

                            log += "\n%s - Error to save questionset %s - %s" % (type_Column.row, sortid, text_en)
                            self.writeLog(log)
                            raise

                    # Type = CATEGORY
                    # Columns required:  Type, Text/Question, Level/Number, Category
                    # Columns optional:  Help text/Description, Slug, Tooltip, Dependencies
                    elif str(type_Column.value) == "Category":
                        self.__handleQuestion(self.CATEGORY, row, type_Column, level_number_column, text_question_Column,
                            _questions_rows, _choices_array, qNumber, questionset, log, _checks, _debug, questionnaire)
                    # Type = QUESTION
                    # Columns required:  Type, Text/Question, Level/Number, Data Type, Category, Stats
                    # Columns optional:  Value List, Help text/Description, Tooltip, Dependencies
                    else:
                        self.__handleQuestion(self.QUESTION, row, type_Column, level_number_column, text_question_Column,
                            _questions_rows, _choices_array, qNumber, questionset, log, _checks, _debug, questionnaire)

        except:
            log += '\nError to save questionsets and questions of the questionnaire %s ' % questionnaire
            self.writeLog(log)
            raise

        log += '\nQuestionnaire %s, questionsets, questions and choices created with success!! ' % questionnaire
        self.writeLog(log)